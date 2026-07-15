import os
import re
import time
from urllib.parse import urljoin, urlparse

import pandas as pd

import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def yes24_crawling(driver, wait, keyword, page_count):
    data = []

    try:
        driver.get("https://www.yes24.com/Main/default.aspx")
        time.sleep(2)

        try:
            search_box = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#query"))
            )
            search_box.clear()
            search_box.send_keys(keyword)
            time.sleep(1)
            search_box.send_keys(Keys.ENTER)
            time.sleep(5)
        except TimeoutException:
            driver.get(
                f"https://www.yes24.com/Product/Search?domain=ALL&query={keyword}"
            )
            time.sleep(3)

        for page in range(1, page_count + 1):
            print(f"{page}페이지 크롤링 중")

            try:
                wait.until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "li[data-goods-no]")
                    )
                )
            except TimeoutException:
                print("검색 결과가 없거나 페이지 로딩이 지연되었습니다.")
                break

            books = driver.find_elements(By.CSS_SELECTOR, "li[data-goods-no]")

            if len(books) == 0:
                print("더 이상 수집할 데이터가 없습니다.")
                break

            before_count = len(data)

            for book in books:
                try:
                    product_links = book.find_elements(
                        By.CSS_SELECTOR,
                        "a[href*='/product/goods/'], a[href*='/Product/Goods/']"
                    )

                    if len(product_links) == 0:
                        continue

                    try:
                        title = book.find_element(By.CSS_SELECTOR, ".gd_name").text.strip()
                    except:
                        title = product_links[0].text.strip()

                    if title == "":
                        continue

                    try:
                        author = book.find_element(By.CSS_SELECTOR, ".info_auth").text.strip()
                    except:
                        author = ""

                    try:
                        price = book.find_element(By.CSS_SELECTOR, ".yes_b").text.strip() + "원"
                    except:
                        price = ""

                    try:
                        publisher = book.find_element(By.CSS_SELECTOR, ".info_pub").text.strip()
                    except:
                        publisher = ""

                    try:
                        publish_date = book.find_element(By.CSS_SELECTOR, ".info_date").text.strip()
                    except:
                        publish_date = ""

                    image_path = ""

                    try:
                        image_tag = book.find_element(By.CSS_SELECTOR, "img")
                        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", image_tag)
                        time.sleep(1)
                        image_url = image_tag.get_attribute("src")
                        
                        if image_url and "noimg" in image_url.lower():
                            image_url = ""
                        
                        if not image_url:
                            image_url = image_tag.get_attribute("data-original")
                        if not image_url:
                            image_url = image_tag.get_attribute("data-src")
                        if not image_url:
                            image_url = image_tag.get_attribute("lazy-src")

                        if image_url:
                            img_response = requests.get(
                                image_url,
                                headers={"User-Agent": "Mozilla/5.0"},
                                timeout=10
                            )

                            extension = os.path.splitext(urlparse(image_url).path)[1]
                            if extension.lower() not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
                                content_type = img_response.headers.get("Content-Type", "")
                                if "png" in content_type:
                                    extension = ".png"
                                elif "gif" in content_type:
                                    extension = ".gif"
                                elif "webp" in content_type:
                                    extension = ".webp"
                                else:
                                    extension = ".jpg"
                                    
                            safe_title = re.sub(r'[\\/:*?"<>|]', "_", title)
                            file_name = f"{len(data) + 1}_{safe_title}{extension}"
                            image_path = os.path.join("images", "yes24", file_name)

                            with open(image_path, "wb") as file:
                                file.write(img_response.content)
                    except Exception as error:
                        print("이미지 저장 실패:", error)

                    data.append({
                        "검색어": keyword,
                        "책제목": title,
                        "저자": author,
                        "가격": price,
                        "출판사": publisher,
                        "출판일": publish_date,
                        "이미지": image_path
                    })

                except Exception as error:
                    print("요소 추출 실패:", error)

            if len(data) == before_count:
                print("현재 페이지에서 수집된 데이터가 없어 종료합니다.")
                break

            if page == page_count:
                break
            
            try:
                next_page = page + 1
                page_links = driver.find_elements(By.CSS_SELECTOR, "a.num")
                clicked = False

                for page_link in page_links:
                    if page_link.text.strip() == str(next_page):
                        driver.execute_script("arguments[0].click();", page_link)
                        clicked = True
                        time.sleep(3)
                        break
            
                if clicked == False:
                    raise Exception("다음 페이지 번호를 찾을 수 없습니다.")
                
            except:
                driver.get(
                    f"https://search.kyobobook.co.kr/search?keyword={keyword}&gbCode=TOT&target=total&page={page + 1}"
                )
                time.sleep(3)
                

    except Exception as error:
        print("크롤링 중 오류 발생:", error)

    return data
 

def kyobo_crawling(driver, wait, keyword, page_count):
    data = []

    try:
        driver.get("https://www.kyobobook.co.kr/")
        time.sleep(2)
        try:
            search_box = wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#searchKeyword"))
            )
            search_box.clear()
            search_box.send_keys(keyword)
            time.sleep(1)
            search_box.send_keys(Keys.ENTER)
            time.sleep(3)

            if keyword not in driver.page_source:
                driver.get(
                    f"https://search.kyobobook.co.kr/search?keyword={keyword}&gbCode=TOT&target=total"
                )
                time.sleep(3)
        except TimeoutException:
            driver.get(
                f"https://search.kyobobook.co.kr/search?keyword={keyword}&gbCode=TOT&target=total"
            )
            time.sleep(3)

        for page in range(1, page_count+1):
            print(f"{page}페이지 크롤링 중")

            try:
                wait.until(EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, ".prod_item")))

            except TimeoutException:
                print("검색 결과가 없거나 페이지 로딩이 지연되었습니다.")
                break

            books = driver.find_elements(By.CSS_SELECTOR, ".prod_item")

            if len(books) == 0:
                print("더 이상 수집할 데이터가 없습니다.")
                break

            before_count = len(data)

            for book in books:
                try:
                    try:
                        title = book.find_element(
                            By.CSS_SELECTOR, "a.prod_info").text.strip()
                    except:
                        title = book.find_element(
                            By.CSS_SELECTOR, "a[href*='product']").text.strip()

                    if title == "":
                        continue

                    try:
                        author = book.find_element(
                            By.CSS_SELECTOR, "a.author").text.strip()
                    except:
                        author = ""

                    try:
                        price = book.find_element(
                            By.CSS_SELECTOR, ".price").text.strip()
                    except:
                        price = ""

                    try:
                        publish = book.find_element(
                            By.CSS_SELECTOR, ".prod_publish").text.strip()
                    except:
                        publish = ""
                    
                    publish_lines = publish.splitlines()
                    publisher = publish_lines[0] if len(publish_lines) > 0 else ""
                    publish_data = publish_lines[1] if len(publish_lines) > 1 else ""
                    
                    image_path = ""
                    
                    try:
                        image_tag = book.find_element(By.CSS_SELECTOR, "img")
                        image_url = image_tag.get_attribute("src")
                        
                        if image_url:
                            image_respons = requests.get(
                                image_url,
                                headers={"User-Agent": "Mozilla/5.0"},
                                timeout=10
                            )
                        extension = os.path.splitext(urlparse(image_url).path)[1] 
                        if extension.lower() not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
                            content_type = image_respons.headers.get("Content-Type", "")
                            if "png" in content_type:
                                extension = ".png"
                            elif "gif" in content_type:
                                extension = ".gif"
                            elif "webp" in content_type:
                                extension = ".webp"
                            else:
                                extension = ".jpg"
                        
                        safe_title = re.sub(r'[\\/:*?"<>|]', "_", title)
                        file_name = f"{len(data) + 1}_{safe_title}{extension}"
                        image_path = os.path.join("images", "kyobo", file_name)
                        
                        with open(image_path, "wb") as file:
                            file.write(image_respons.content)
                    
                    except Exception as error:
                        print("이미지 저장 실패", error)
                    
                    data.append({
                        "검색어": keyword,
                        "책제목": title,
                        "저자": author,
                        "가격": price,
                        "출판사": publisher,
                        "출판일": publish_data,
                        "이미지": image_path
                    })

                except Exception as error:
                    print("요소 추출 실패", error)
            
            if len(data) == before_count:
                print("현재 페이지에서 수집된 데이터가 없어 종료합니다.")
                break
        
            if page == page_count:
                break
            
            try:
                next_page = page + 1
                page_links = driver.find_elements(By.CSS_SELECTOR, "a.btn_page_num")
                clicked = False

                for page_link in page_links:
                    if page_link.text.strip() == str(next_page):
                        driver.execute_script("arguments[0].click();", page_link)
                        clicked = True
                        time.sleep(3)
                        break

                if clicked == False:
                    raise Exception("다음 페이지 번호를 찾을 수 없습니다.")
            except:
                driver.get(
                    f"https://search.kyobobook.co.kr/search?keyword={keyword}&gbCode=TOT&target=total&page={page + 1}"
                )
                time.sleep(3)
                
                
    except Exception as error:
        print("크롤링 중 오류 발생", error)
    
    return data

def aladin_crawling(driver, wait, keyword, page_count):
    data = []
        
    try:
        driver.get("https://www.aladin.co.kr/home/welcome.aspx")
        time.sleep(2)
        
        try:
            search_box = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#SearchWord"))
            )
            search_box.send_keys(keyword)
            time.sleep(1)
            search_box.send_keys(Keys.ENTER)
            time.sleep(3)
            
        except TimeoutException:
            driver.get(
                f"https://www.aladin.co.kr/search/wsearchresult.aspx?SearchTarget=All&SearchWord={keyword}"
            )
            time.sleep(3)
        
        for page in range(1, page_count+1):
            print(f"{page}페이지 크롤링 중")
            
            try:
                wait.until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".ss_book_box"))
                )
            except TimeoutException:
                print("검색 결과가 없거나 페이지 로딩이 지연되었습니다.")
                break
                
            books = driver.find_elements(By.CSS_SELECTOR, ".ss_book_box")
            
            if len(books) == 0:
                print("더 이상 수집할 데이터가 없습니다.")
                break
            
            before_count = len(data)
            
            for book in books:
                try:
                    title = ""

                    title_tag = book.find_elements(By.CSS_SELECTOR, ".bo3")
                    if len(title_tag) > 0:
                        title = title_tag[0].text.strip()
                        
                    if title == "":
                        title_tag = book.find_elements(By.CSS_SELECTOR, "a[herf*='ItemId']")
                        if len(title_tag) > 0:
                            title = title_tag[0].text.strip()
                    
                    if title == "":
                        continue       
                                       
                    try:
                        info_items = book.find_elements(By.CSS_SELECTOR, ".ss_book_list li")

                        for item in info_items:
                            info_line = item.text.strip()
                            info_parts = [
                                part.strip()
                                for part in info_line.split("|")
                                if part.strip()
                            ]

                            if len(info_parts) >= 3:
                                author = info_parts[0]
                                publisher = info_parts[1]
                                publish_date = info_parts[2]
                                break
                    except:
                        author = ""
                        publisher = ""
                        publish_date = ""
                                                   
                    try:
                        price = book.find_element(By.CSS_SELECTOR, ".ss_p2").text.strip()
                    except:
                        price = ""
                    
                    image_path = ""
                    
                    try:
                        image_tags = book.find_elements(By.CSS_SELECTOR, "img.front_cover")
                        
                        if len(image_tags) == 0:
                            image_tags = book.find_elements(By.CSS_SELECTOR, "img.i_cover")
                        
                        if len(image_tags) == 0:
                            image_tags = book.find_elements(By.CSS_SELECTOR, "img")
                        
                        if len(image_tags) == 0:
                            image_url = ""
                        else:
                            image_tag = image_tags[0]
                            image_url = image_tag.get_attribute("src")
                        
                        if image_url:
                            image_url = urljoin("https://www.aladin.co.kr", image_url)
                            
                            image_response = requests.get(
                                image_url,
                                headers={"User-Agent": "Mozilla/5.0"},
                                timeout=10
                            )
                            extension = os.path.splitext(urlparse(image_url).path)[1]
                            if extension.lower() not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
                                content_type = image_response.headers.get("Content-Type", "")
                                if "png" in content_type:
                                    extension = ".png"
                                elif "gif" in content_type:
                                    extension = ".gif"
                                elif "webp" in content_type:
                                    extension = ".webp"
                                else:
                                    extension = ".jpg"
                            
                            safe_title = re.sub(r'[\\/:*?"<>|]', "_", title)
                            file_name = f"{len(data)+1}_{safe_title}{extension}"
                            image_path = os.path.join("images", "aladin", file_name)
                            
                            with open(image_path, "wb") as file:
                                file.write(image_response.content)
                    except Exception as error:
                        print("이미지 저장 실패", error)
                        break
                    
                    data.append({
                        "검색어": keyword,
                        "책제목": title,
                        "저자": author,
                        "가격": price,
                        "출판사": publisher,
                        "출판일": publish_date,
                        "이미지": image_path
                    })
                except Exception as error:
                    print("요소 추출 실패", error)
            
            if len(data) == before_count:
                print("현재 페이지에서 수집된 데이터가 없어 종료합니다.")
                break
            
            if page == page_count:
                break
            
            try:
                next_page = page + 1
                page_links = driver.find_elements(By.CSS_SELECTOR, "a.numoff")
                clicked = False

                for page_link in page_links:
                    if page_link.text.strip() == str(next_page):
                        driver.execute_script("arguments[0].click();", page_link)
                        clicked = True
                        time.sleep(3)
                        break

                if clicked == False:
                    raise Exception("다음 페이지 번호를 찾을 수 없습니다.")
            except:
                driver.get(
                    f"https://www.aladin.co.kr/search/wsearchresult.aspx?SearchTarget=All&SearchWord={keyword}&page={page + 1}"
                )
                time.sleep(3)

    except Exception as error:
        print("크롤링 중 오류 발생:", error)

    return data


def crawl_start(keyword, yes24_page, kyobo_page, aladin_page):
    start_time = time.time()
    
    driver = webdriver.Chrome()
    driver.maximize_window()
    wait = WebDriverWait(driver, 10)

    os.makedirs("images", exist_ok=True)
    os.makedirs("images/yes24", exist_ok=True)
    os.makedirs("images/kyobo", exist_ok=True)
    os.makedirs("images/aladin", exist_ok=True)

    try:
        yes24_data = yes24_crawling(driver, wait, keyword, yes24_page)
        kyobo_data = kyobo_crawling(driver, wait, keyword, kyobo_page)
        aladin_data = aladin_crawling(driver, wait, keyword, aladin_page)

        columns = ["검색어", "책제목", "저자", "가격", "출판사", "출판일", "이미지"]

        yes24_df = pd.DataFrame(yes24_data, columns=columns)
        kyobo_df = pd.DataFrame(kyobo_data, columns=columns)
        aladin_df = pd.DataFrame(aladin_data, columns=columns)

        file_name = "book_crawling.xlsx"

        with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
            yes24_df.to_excel(writer, sheet_name="yes24", index=False)
            kyobo_df.to_excel(writer, sheet_name="kyobo", index=False)
            aladin_df.to_excel(writer, sheet_name="aladin", index=False)
            
        workbook = load_workbook(file_name)
        
        for sheet_name in ["yes24", "kyobo", "aladin"]:
            worksheet = workbook[sheet_name]
            
            for column in ["A", "B", "C", "D", "E", "F"]:
                max_length = 0
                
                for cell in worksheet[column]:
                    if cell.value:
                        cell_length = 0
                        
                        for text in str(cell.value):
                            if ord(text) > 127:
                                cell_length += 2
                            else:
                                cell_length += 1
                        
                        if cell_length > max_length:
                            max_length = cell_length
                
                if column == "A":
                    worksheet.column_dimensions[column].width = max(max_length + 2, 12)
                else:
                    worksheet.column_dimensions[column].width = max_length + 2
            
            worksheet.column_dimensions["G"].width = 18
            
            for row in range(2, worksheet.max_row + 1):
                image_path = worksheet[f"G{row}"].value
                
                if image_path and os.path.exists(image_path):
                    image = Image(image_path)
                    image.width = 80
                    image.height = 110
                    
                    worksheet[f"G{row}"].value = ""
                    worksheet.row_dimensions[row].height = 85
                    worksheet.add_image(image, f"G{row}")
        
        workbook.save(file_name)

        end_time = time.time()

        print("엑셀 저장 완료:", file_name)
        print("YES24 데이터 수집 개수:", len(yes24_df))
        print("교보문고 데이터 수집 개수:", len(kyobo_df))
        print("알라딘 데이터 수집 개수:", len(aladin_df))
        print(f"전체 크롤링 시간: {end_time - start_time:.2f}초")
    
    finally:
        driver.quit()


crawl_start("헤르만 헤세", 1, 2, 1)
